"""
table_EA_comparison.py
======================
Produce un Excel con tutte le regioni DK (surface area × EA) affiancando:
  - colonne Grasby et al. 2020 (rg, SE, p, hit)
  - colonne This study BIG40  (rg, SE, p, hit)

Riusa la stessa logica di fig_replication_scatter.py.

MODIFICA LE 3 PATH PRIMA DI RUNNARE.
"""

import os
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── PATH (modifica qui) ───────────────────────────────────────────────────────
RG_CSV      = os.path.expanduser("~/ldsc_project/rg_ALL_163_IDPs_FINAL.csv")
GRASBY_XLSX = os.path.expanduser("~/ldsc_project/grasby_tables.xlsx")
OUT_XLSX    = os.path.expanduser("~/ldsc_project/results/Table_EA_comparison.xlsx")
# ─────────────────────────────────────────────────────────────────────────────

# ── BIG40 IDP → (dk_name, hemisphere) ────────────────────────────────────────
BIG40_MAP = {
    "0649": ("bankssts",                "lh"),
    "0650": ("caudalanteriorcingulate", "lh"),
    "0651": ("caudalmiddlefrontal",     "lh"),
    "0652": ("cuneus",                  "lh"),
    "0653": ("entorhinal",              "lh"),
    "0654": ("fusiform",                "lh"),
    "0655": ("inferiorparietal",        "lh"),
    "0656": ("inferiortemporal",        "lh"),
    "0657": ("isthmuscingulate",        "lh"),
    "0658": ("lateraloccipital",        "lh"),
    "0659": ("lateralorbitofrontal",    "lh"),
    "0660": ("lingual",                 "lh"),
    "0661": ("medialorbitofrontal",     "lh"),
    "0662": ("middletemporal",          "lh"),
    "0663": ("parahippocampal",         "lh"),
    "0664": ("paracentral",             "lh"),
    "0665": ("parsopercularis",         "lh"),
    "0666": ("parsorbitalis",           "lh"),
    "0667": ("parstriangularis",        "lh"),
    "0668": ("pericalcarine",           "lh"),
    "0669": ("postcentral",             "lh"),
    "0670": ("posteriorcingulate",      "lh"),
    "0671": ("precentral",              "lh"),
    "0672": ("precuneus",               "lh"),
    "0673": ("rostralanteriorcingulate","lh"),
    "0674": ("rostralmiddlefrontal",    "lh"),
    "0675": ("superiorfrontal",         "lh"),
    "0676": ("superiorparietal",        "lh"),
    "0677": ("superiortemporal",        "lh"),
    "0678": ("supramarginal",           "lh"),
    "0679": ("frontalpole",             "lh"),
    "0680": ("transversetemporal",      "lh"),
    "0681": ("insula",                  "lh"),
    "0683": ("bankssts",                "rh"),
    "0684": ("caudalanteriorcingulate", "rh"),
    "0685": ("caudalmiddlefrontal",     "rh"),
    "0686": ("cuneus",                  "rh"),
    "0687": ("entorhinal",              "rh"),
    "0688": ("fusiform",                "rh"),
    "0689": ("inferiorparietal",        "rh"),
    "0690": ("inferiortemporal",        "rh"),
    "0691": ("isthmuscingulate",        "rh"),
    "0692": ("lateraloccipital",        "rh"),
    "0693": ("lateralorbitofrontal",    "rh"),
    "0694": ("lingual",                 "rh"),
    "0695": ("medialorbitofrontal",     "rh"),
    "0696": ("middletemporal",          "rh"),
    "0697": ("parahippocampal",         "rh"),
    "0698": ("paracentral",             "rh"),
    "0699": ("parsopercularis",         "rh"),
    "0700": ("parsorbitalis",           "rh"),
    "0701": ("parstriangularis",        "rh"),
    "0702": ("pericalcarine",           "rh"),
    "0703": ("postcentral",             "rh"),
    "0704": ("posteriorcingulate",      "rh"),
    "0705": ("precentral",              "rh"),
    "0706": ("precuneus",               "rh"),
    "0707": ("rostralanteriorcingulate","rh"),
    "0708": ("rostralmiddlefrontal",    "rh"),
    "0709": ("superiorfrontal",         "rh"),
    "0710": ("superiorparietal",        "rh"),
    "0711": ("superiortemporal",        "rh"),
    "0712": ("supramarginal",           "rh"),
    "0713": ("frontalpole",             "rh"),
    "0714": ("transversetemporal",      "rh"),
    "0715": ("insula",                  "rh"),
}

DK_TO_GRASBY = {
    "bankssts":                "Banks of the Superior Temporal Sulcus",
    "caudalanteriorcingulate": "Caudal Anterior Cingulate",
    "caudalmiddlefrontal":     "Caudal Middle Frontal",
    "cuneus":                  "Cuneus",
    "entorhinal":              "Entorhinal",
    "fusiform":                "Fusiform",
    "inferiorparietal":        "Inferior Parietal",
    "inferiortemporal":        "Inferior Temporal",
    "isthmuscingulate":        "Isthmus Cingulate",
    "lateraloccipital":        "Lateral Occipital",
    "lateralorbitofrontal":    "Lateral Orbitofrontal",
    "lingual":                 "Lingual",
    "medialorbitofrontal":     "Medial Orbitofrontal",
    "middletemporal":          "Middle Temporal",
    "parahippocampal":         "Parahippocampal",
    "paracentral":             "Paracentral",
    "parsopercularis":         "Pars Opercularis",
    "parsorbitalis":           "Pars Orbitalis",
    "parstriangularis":        "Pars Triangularis",
    "pericalcarine":           "Pericalcarine",
    "postcentral":             "Postcentral",
    "posteriorcingulate":      "Posterior Cingulate",
    "precentral":              "Precentral",
    "precuneus":               "Precuneus",
    "rostralanteriorcingulate":"Rostral Anterior Cingulate",
    "rostralmiddlefrontal":    "Rostral Middle Frontal",
    "superiorfrontal":         "Superior Frontal",
    "superiorparietal":        "Superior Parietal",
    "superiortemporal":        "Superior Temporal",
    "supramarginal":           "Supramarginal",
    "frontalpole":             "Frontal Pole",
    "transversetemporal":      "Transverse Temporal",
    "insula":                  "Insula",
}

# ── 1. Carica tuoi dati ───────────────────────────────────────────────────────
print("Leggo rg_ALL_163_IDPs_FINAL.csv ...")
rg_raw = pd.read_csv(RG_CSV)
rg_raw["IDP"] = rg_raw["IDP"].astype(str).str.zfill(4)

rows = []
for _, row in rg_raw.iterrows():
    idp = row["IDP"]
    if idp not in BIG40_MAP:
        continue
    dk_name, hemi = BIG40_MAP[idp]
    rows.append({"dk_name": dk_name, "hemi": hemi,
                 "rg": row["rg"], "rg_se": row["rg_se"], "rg_p": row["rg_p"]})

df_big40 = pd.DataFrame(rows)

df_mine = (
    df_big40
    .groupby("dk_name")
    .agg(rg=("rg","mean"), rg_se=("rg_se","mean"), rg_p=("rg_p","mean"))
    .reset_index()
)
df_mine["grasby_name"] = df_mine["dk_name"].map(DK_TO_GRASBY)
print(f"  Regioni tuo dataset: {len(df_mine)}")

# ── 2. Carica Grasby S17 ──────────────────────────────────────────────────────
print("Leggo Grasby S17 ...")
wb_in = openpyxl.load_workbook(GRASBY_XLSX, read_only=True)
ws_in = wb_in["Table S17"]
rows_excel = list(ws_in.iter_rows(values_only=True))

EA_RG_COL, EA_SE_COL, EA_P_COL = 33, 34, 35

grasby_rows = []
for row in rows_excel[5:]:
    if row[0] is None:
        continue
    if row[1] != "SA":
        continue
    if row[2] == "Total Surface Area":
        continue
    grasby_rows.append({
        "grasby_name": row[2],
        "rg_grasby":   row[EA_RG_COL],
        "se_grasby":   row[EA_SE_COL],
        "p_grasby":    row[EA_P_COL],
    })

df_grasby = pd.DataFrame(grasby_rows)
print(f"  Regioni Grasby SA: {len(df_grasby)}")

# ── 3. Merge ──────────────────────────────────────────────────────────────────
merged = df_mine.merge(df_grasby, on="grasby_name", how="inner")
merged["hit_grasby"] = merged["p_grasby"] < 0.05
merged["hit_mine"]   = merged["rg_p"]     < 0.05

# Sort: prima i "both hit", poi "solo grasby", poi "solo mine", poi nessuno
def sort_key(row):
    if row["hit_grasby"] and row["hit_mine"]:   return 0
    if row["hit_grasby"] and not row["hit_mine"]:return 1
    if not row["hit_grasby"] and row["hit_mine"]:return 2
    return 3

merged["_sort"] = merged.apply(sort_key, axis=1)
merged = merged.sort_values(["_sort", "grasby_name"]).drop(columns="_sort").reset_index(drop=True)

print(f"  Regioni totali: {len(merged)}")
print(f"  Hit Grasby: {merged['hit_grasby'].sum()}")
print(f"  Hit Mine:   {merged['hit_mine'].sum()}")
print(f"  Hit entrambi: {(merged['hit_grasby'] & merged['hit_mine']).sum()}")

# ── 4. Scrivi Excel ───────────────────────────────────────────────────────────
os.makedirs(os.path.dirname(OUT_XLSX), exist_ok=True)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "EA_comparison"

# ── Stili ─────────────────────────────────────────────────────────────────────
FILL_HEADER_G  = PatternFill("solid", fgColor="1F4E79")   # blu scuro Grasby
FILL_HEADER_M  = PatternFill("solid", fgColor="833C00")   # arancio scuro Mine
FILL_HIT_G     = PatternFill("solid", fgColor="BDD7EE")   # azzurro chiaro
FILL_HIT_M     = PatternFill("solid", fgColor="FCE4D6")   # arancio chiaro
FILL_BOTH      = PatternFill("solid", fgColor="E2EFDA")   # verde hit entrambi
FILL_SUBHDR    = PatternFill("solid", fgColor="D9D9D9")   # grigio sub-header

FONT_HDR   = Font(bold=True, color="FFFFFF", size=10)
FONT_SUBHDR= Font(bold=True, size=9)
FONT_BODY  = Font(size=9)
FONT_HIT   = Font(bold=True, size=9)

ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left",   vertical="center")
ALIGN_R = Alignment(horizontal="right",  vertical="center")

thin = Side(style="thin", color="AAAAAA")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

FMT_NUM = "0.0000"
FMT_P   = "0.0000"

# ── Riga 1: gruppo header ─────────────────────────────────────────────────────
ws.merge_cells("A1:A2")
ws["A1"] = "Brain Region"
ws["A1"].fill = PatternFill("solid", fgColor="404040")
ws["A1"].font = Font(bold=True, color="FFFFFF", size=10)
ws["A1"].alignment = ALIGN_C

ws.merge_cells("B1:E1")
ws["B1"] = "Grasby et al. 2020"
ws["B1"].fill = FILL_HEADER_G
ws["B1"].font = FONT_HDR
ws["B1"].alignment = ALIGN_C

ws.merge_cells("F1:I1")
ws["F1"] = "This Study (BIG40)"
ws["F1"].fill = FILL_HEADER_M
ws["F1"].font = FONT_HDR
ws["F1"].alignment = ALIGN_C

# ── Riga 2: sub-header ────────────────────────────────────────────────────────
sub_headers = ["rg", "SE", "p", "Hit (p<0.05)",
               "rg", "SE", "p", "Hit (p<0.05)"]
for col_idx, label in enumerate(sub_headers, start=2):
    cell = ws.cell(row=2, column=col_idx, value=label)
    cell.fill = FILL_SUBHDR
    cell.font = FONT_SUBHDR
    cell.alignment = ALIGN_C
    cell.border = BORDER

ws["A2"].fill = FILL_SUBHDR
ws["A2"].border = BORDER

# ── Dati ─────────────────────────────────────────────────────────────────────
for r_idx, row_data in merged.iterrows():
    excel_row = r_idx + 3   # dati partono da riga 3

    both = row_data["hit_grasby"] and row_data["hit_mine"]
    fill_row_g = FILL_BOTH if both else (FILL_HIT_G if row_data["hit_grasby"] else None)
    fill_row_m = FILL_BOTH if both else (FILL_HIT_M if row_data["hit_mine"]   else None)

    # Colonna A: nome regione
    c = ws.cell(row=excel_row, column=1, value=row_data["grasby_name"])
    c.font = FONT_BODY
    c.alignment = ALIGN_L
    c.border = BORDER
    if both:
        c.fill = FILL_BOTH

    # Colonne B-E: Grasby
    grasby_vals = [
        row_data["rg_grasby"],
        row_data["se_grasby"],
        row_data["p_grasby"],
        "Yes" if row_data["hit_grasby"] else "No",
    ]
    for col_offset, val in enumerate(grasby_vals):
        c = ws.cell(row=excel_row, column=2 + col_offset, value=val)
        c.alignment = ALIGN_R
        c.border = BORDER
        if fill_row_g:
            c.fill = fill_row_g
        if row_data["hit_grasby"]:
            c.font = FONT_HIT
        else:
            c.font = FONT_BODY
        if isinstance(val, float):
            c.number_format = FMT_NUM

    # Colonne F-I: Mine
    mine_vals = [
        row_data["rg"],
        row_data["rg_se"],
        row_data["rg_p"],
        "Yes" if row_data["hit_mine"] else "No",
    ]
    for col_offset, val in enumerate(mine_vals):
        c = ws.cell(row=excel_row, column=6 + col_offset, value=val)
        c.alignment = ALIGN_R
        c.border = BORDER
        if fill_row_m:
            c.fill = fill_row_m
        if row_data["hit_mine"]:
            c.font = FONT_HIT
        else:
            c.font = FONT_BODY
        if isinstance(val, float):
            c.number_format = FMT_NUM

# ── Larghezze colonne ─────────────────────────────────────────────────────────
col_widths = [34, 9, 9, 9, 12,   9, 9, 9, 12]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Freeze top 2 righe ────────────────────────────────────────────────────────
ws.freeze_panes = "A3"

# ── Nota in fondo ─────────────────────────────────────────────────────────────
note_row = len(merged) + 4
ws.cell(row=note_row, column=1,
        value=("Note: BIG40 rg = mean of left and right hemisphere LDSC estimates. "
               "Hit = p < 0.05 (nominal). "
               "Temporal Pole excluded (not available in BIG40). "
               "Grasby SE/p from Table S17 (surface area × EA, pooled hemisphere). "
               "Rows sorted: both hits → Grasby-only hit → Mine-only hit → neither."))
ws.cell(row=note_row, column=1).font = Font(italic=True, size=8, color="666666")
ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)
ws.merge_cells(start_row=note_row, start_column=1,
               end_row=note_row, end_column=9)
ws.row_dimensions[note_row].height = 40

wb.save(OUT_XLSX)
print(f"\nExcel salvato: {OUT_XLSX}")
print(f"Righe dati: {len(merged)}  |  Hit Grasby: {merged['hit_grasby'].sum()}  |  Hit Mine: {merged['hit_mine'].sum()}")
